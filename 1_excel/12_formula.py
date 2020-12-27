import datetime
from openpyxl import Workbook
wb = Workbook()
ws = wb.active

ws["A1"] = datetime.datetime.today()
ws["A2"] = "=SUM(1,2,3)"
ws["A3"] = "=AVERAGE(1,2,3)"

wb.save("sample_formula.xlsx")