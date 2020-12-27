from openpyxl import Workbook
from openpyxl.utils.cell import coordinate_from_string
wb = Workbook()
ws = wb.active

datas=[ 
("학번", "출석(10)", "퀴즈1(10)", "퀴즈2(10)", "중간고사(20)", "기말고사(30)", "프로젝트(20)","총점"),
(1,10,8,5,14,26,12),
(2,7,3,7,15,24,18),
(3,9,5,8,8,12,4),
(4,7,8,7,17,21,18),
(5,7,8,7,16,25,15),
(6,3,5,8,8,17,0),
(7,4,9,10,16,27,18),
(8,6,6,6,15,19,17),
(9,10,10,9,19,30,19),
(10,9,8,8,20,25,20)]

for s in datas:
    ws.append(s)

for idx, cell in enumerate(ws["D"]):
    if idx == 0:
        continue
    cell.value = 10

for y in range(1,ws.max_row + 1):
    if(y==1):
        ws.cell(row=y, column=8, value="총점")
    else:
        ws.cell(row=y, column=8, value="=SUM(B{a}:G{a})".format(a=y))
        result = 0
        for x in range(1,8):
            if(x==1):
                ws.cell(row=x, column=9, value="성적")
            else:
                result += int(ws.cell(row=y, column=x).value)
                
                attend = int(ws.cell(row=y, column=2).value)
                if attend < 5:
                    ws.cell(row=y, column=9, value="F")
                elif result >= 90:
                    ws.cell(row=y, column=9, value="A")
                elif result >= 80:
                    ws.cell(row=y, column=9, value="B")
                elif result >= 70:
                    ws.cell(row=y, column=9, value="C")
                else: 
                    ws.cell(row=y, column=9, value="D")

wb.save("./python_rpa_basic/1_excel/quiz.xlsx")

