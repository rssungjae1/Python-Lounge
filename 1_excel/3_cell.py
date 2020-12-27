from openpyxl import Workbook
wb = Workbook() 

ws = wb.active
ws.title = "SongSheet"

ws["A1"] = 1
ws["A2"] = 2
ws["A3"] = 3

ws["B1"] = 4
ws["B2"] = 5
ws["B3"] = 6

print(ws["A1"])
print(ws["A1"].value) # 값 출력
print(ws["A10"].value) # 값이 없을땐 None 출력
print(ws.cell(row=1, column=1).value) # ws["A1"].value
print(ws.cell(row=1, column=2).value) # ws["B1"].value

c = ws.cell(column=3, row=1, value=10)
print(c.value)

from random import *
index = 1
for x in range(1, 11):
    for y in  range(1,11):
        #ws.cell(row=x, column=y, value=randint(1, 100))
        ws.cell(row=x, column=y, value=index)
        index+=1
wb.save("sample.xlsx") # 워크북 저장
wb.close()