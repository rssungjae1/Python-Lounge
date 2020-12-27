from openpyxl import load_workbook
# wb = load_workbook("sample_formula.xlsx")
# ws = wb.active
# 수식 그대로 가져오기
# for row in ws.values:
#     for cell in row:
#         print(cell)

wb = load_workbook("sample_formula.xlsx", data_only=True)
ws = wb.active

# 수식이 아닌 실제 데이터 가져오기
# evaluate 되지 않은 상태의 데이터는 None
# 엑셀 파일을 다시 열었다가 다시 저장하고 하는걸로 ㄱ
# a파일 열고 저장하고 a파일을 가지고 b파일을 가지고 수식연산하려면 불가
# 사람이 열고 저장해야한다.
for row in ws.values:
    for cell in row:
        print(cell)