from openpyxl import Workbook, load_workbook
# wb = Workbook()
# ws = wb.active

# ws.merge_cells("B2:D2")
# ws["B2"].value= "Merged.Cell"

# wb.save("sample_merge.xlsx")

wb = load_workbook("sample_merge.xlsx")
ws = wb.active

ws.unmerge_cells("B2:D2")
wb.save("sample_unmerge.xlsx")