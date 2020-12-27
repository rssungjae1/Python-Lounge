from openpyxl import Workbook

wb = Workbook() # 새 워크북 생성
# ws = wb.active # 현재 활성화 된 sheet 가져옴

ws = wb.create_sheet()
ws.title = "MySheet" # 워크시트 이름변경
ws.sheet_properties.tabColor = "ccffff" # RGB형태로 값을 넣어 색 변경

# Sheet, MySheet, YourSheet
ws1 = wb.create_sheet("sheet1")
ws2 = wb.create_sheet("sheet2",2) # 2번째 index에 sheet 생성

new_ws = wb["sheet2"] # dict형태로 sheet에 접근가능

print(wb.sheetnames) # 모든 sheet 이름 확인

# sheet 복사
new_ws["A1"] = "test"
target = wb.copy_worksheet(new_ws)
target.title = " Copied Sheet"

print(wb.sheetnames) # 모든 sheet 이름 확인

wb.save("sample.xlsx") # 워크북 저장
wb.close()